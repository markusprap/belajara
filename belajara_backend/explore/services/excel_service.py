import openpyxl

def extract_text_from_excel(excel_file) -> str:
    """
    Extracts all text from an uploaded Excel file in-memory.
    Iterates through sheets and formats rows in a clean tabular string format.
    """
    try:
        wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        text_parts = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text_parts.append(f"Sheet: {sheet_name}\n")
            for row in sheet.iter_rows(values_only=True):
                # Filter out completely empty rows and format cell values
                row_str = " | ".join(str(cell).strip() for cell in row if cell is not None)
                if row_str.strip():
                    text_parts.append(row_str)
            text_parts.append("\n" + "="*40 + "\n")
        return "\n".join(text_parts).strip()
    except Exception as e:
        raise ValueError(f"Gagal memproses file Excel: {str(e)}")
